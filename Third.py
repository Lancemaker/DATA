import os
from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl
from lib.Carro import Carro
import re

#variaveis de arquivo
raizArquivo = os.getcwd()+"/"
print(raizArquivo)
planilhasArquivo = raizArquivo+"planilhas/"
saida = raizArquivo+"saida/"
diagramar = "Planilha_Diagramar_Completa_v03.xlsx"
completa = "Planilha AUTOMÓVEIS REV_240918_v2.xlsx"
basica= "planilha.xlsx"


workbook= load_workbook(planilhasArquivo+basica)

planilha=workbook.get_sheet_by_name('Filtro_3')
TabFiltro4=workbook.create_sheet("Filtro_4")

carros = []
carrosFiltrados=[]

for num in range(2,2045):
    tempString=""
    carro = Carro(planilha['A'+str(num)].value,planilha['B'+str(num)].value,planilha['C'+str(num)].value,planilha['D'+str(num)].value.split(" "),planilha['E'+str(num)].value,planilha['F'+str(num)].value)
    
    if(carro.montadora != "LAND ROVER")or(carro.montadora != "TOYOTA"):
        carro.ano=carro.ano[0]
        if(carro.ano[-1]==" " or carro.ano[-1]==''):
            carro.ano=carro.ano[:-1]
        #tratando data

    i=0
    while i<len(carro.ano):
        carro.ano[i]=carro.ano[i].replace('[','')
        carro.ano[i]=carro.ano[i].replace(']','')
        carro.ano[i]=carro.ano[i].replace(',','')
        #carro.ano[i]=carro.ano[i].replace('\'','')
        i+=1
    #Regras para renomear os carros:
    if(len(carros)==0):
        carros.append(carro)
    else:
        last=carros[-1]

        if(carro.montadora==last.montadora and carro.modelo==last.modelo and carro.tipo==last.tipo and carro.capacidade==last.capacidade and carro.recomendacao==last.recomendacao):
            last.ano=last.ano+carro.ano
            
            '''x=y=0
            while x<len(carro.ano):                
                while y<len(last.ano):                    
                    if carro.ano[x] not in last.ano:
                        last.ano.append(carro.ano[x])
                        if(not carro.ano[x].isalpha()):
                            last.ano.sort()    
                                           
                    y+=1  
                x+=1
            '''
        else:                 
            carros.append(carro)     
        print(carros[-1].modelo)   
#for c in carros:
    #print(c.montadora,c.modelo,c.tipo,c.ano,c.capacidade,c.recomendacao)
#print(len(carros))



#gerando a planilha :D
def addRotulos(tabela):
        tabela['A1']="MONTADORA"
        tabela['B1']="MODELO"
        tabela['C1']="TIPO"
        tabela['D1']="ANO"
        tabela['E1']="CAPACIDADE"
        tabela['F1']="RECOMENDACAO_CASTROL"


addRotulos(TabFiltro4)
count=1
for c in carros:
    t=""
    count+=1
    TabFiltro4['A'+str(count)]=c.montadora
    TabFiltro4['B'+str(count)]=c.modelo
    TabFiltro4['C'+str(count)]=c.tipo
    if(c.montadora != 'TOYOTA'):
        print(c.montadora)
        if(len(c.ano)>1):
            c.ano.sort()
            c.ano=[c.ano[0]+" ate "+c.ano[-1]]
        TabFiltro4['D'+str(count)]=c.ano[0]
    else:
        
        if(c.montadora=='TOYOTA'):
            x=0
            t=""
            while(x<len(c.ano)):
                c.ano[x]=c.ano[x].replace('[','')
                c.ano[x]=c.ano[x].replace(']','')
                c.ano[x]=c.ano[x].replace('\"','')
                c.ano[x]=c.ano[x].replace('\'','')
                c.ano[x]=c.ano[x].replace(',','')    
                t=t+" "+c.ano[x]
                x+=1                       
        TabFiltro4['D'+str(count)]=t
    TabFiltro4['E'+str(count)]=c.capacidade
    TabFiltro4['F'+str(count)]=c.recomendacao
workbook.save('planilha.xlsx')

