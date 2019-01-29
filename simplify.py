import os
from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl
from lib.Carro import Carro
import re

#variaveis de arquivo
raizArquivo = os.getcwd()+"/"
print(raizArquivo)
planilhasArquivo = raizArquivo+"planilhas/"
saida = raizArquivo+"saida/"
diagramar = "Planilha_Diagramar_Completa_v03.xlsx"
completa = "Planilha AUTOMÓVEIS REV_240918_v2.xlsx"
basica= "planilha.xlsx"


workbook= load_workbook(planilhasArquivo+basica)

planilha=workbook.get_sheet_by_name('Filtro_2')
TabFiltro3=workbook.create_sheet("Filtro_3")

carros = []
carrosFiltrados=[]

for num in range(2,2923):
    tempString=""
    carro = Carro(planilha['A'+str(num)].value,planilha['B'+str(num)].value,planilha['C'+str(num)].value,planilha['D'+str(num)].value.split(" "),planilha['E'+str(num)].value,planilha['F'+str(num)].value)
    
    carro.ano=carro.ano[0]
    carro.ano=carro.ano[:-1]
    #Regras para renomear os carros:
    temp=carro.modelo.split(" ")
    
    if  ("DPF" in temp):
        for item in temp:
            tempString+=item+" "
        carro.modelo=tempString
    else:
        carro.modelo=temp[0]
    
    print(num,carro.modelo,carro.ano)
     

    #com erro linha 43
    #print(num, carrosFiltrados[num-3].montadora, carrosFiltrados[num-3].modelo, carrosFiltrados[num-3].tipo, carrosFiltrados[num-3].ano, carrosFiltrados[num-3].capacidade, carrosFiltrados[num-3].recomendacao)