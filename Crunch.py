# Import `os` 
import os
from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl
from lib.Carro import Carro
import re


#variaveis de arquivo
raizArquivo = os.getcwd()+"/"
print(raizArquivo)
planilhasArquivo = raizArquivo+"planilhas/"
saida = raizArquivo+"saida/"
diagramar = "Planilha_Diagramar_Completa_v03.xlsx"
completa = "Planilha AUTOMÓVEIS REV_240918_v2.xlsx"
basica= "basica.xlsx"

workbook= load_workbook(planilhasArquivo+basica)
#Para pegar as worksheets(aba das tabelas) e mostrar a tabela ativa
#print(workbook.sheetnames , workbook.active)

#MONTADORA 	MODELO 	TIPO	 ANO 	CAPACIDADE	RECOMENDACAO_CASTROL
planilha=workbook.get_sheet_by_name("Planilha1")

TabFiltro1=workbook.create_sheet("Filtro_1")
TabFiltro2=workbook.create_sheet("Filtro_2")

carros = []
carrosFiltrados =[]
filtroModelo=[]
for num in range(2,7833):
#preenche os objetos
    carro = Carro(planilha['A'+str(num)].value,planilha['B'+str(num)].value,planilha['C'+str(num)].value,planilha['D'+str(num)].value,planilha['E'+str(num)].value,planilha['F'+str(num)].value)    
    print(num)
    catch = re.findall(r"[-+]?\d*\.\d", carro.modelo)
    if(len(catch)>0):
        carro.tipo = catch[0] 
        carro.modelo = carro.modelo.replace(carro.tipo+" ","")
    else:
        carro.tipo = "NAO ENCONTRADO"  
    
    carros.append(carro)   
    if len(carrosFiltrados)==0:
        carrosFiltrados.append(carro)
    else:
        if(carrosFiltrados[-1].montadora==carro.montadora and carrosFiltrados[-1].modelo==carro.modelo and carrosFiltrados[-1].tipo==carro.tipo and carrosFiltrados[-1].recomendacao==carro.recomendacao and carrosFiltrados[-1].capacidade==carro.capacidade):
            if carro.ano not in carrosFiltrados[-1].ano:
                carrosFiltrados[-1].ano.append(carro.ano[0])
        else:
            carrosFiltrados.append(carro)


#Rotulos
def addRotulos(tabela):
        tabela['A1']="MONTADORA"
        tabela['B1']="MODELO"
        tabela['C1']="TIPO"
        tabela['D1']="ANO"
        tabela['E1']="CAPACIDADE"
        tabela['F1']="RECOMENDACAO_CASTROL"


addRotulos(TabFiltro1)
count=1

for c in carrosFiltrados:
    t=""
    count+=1
    TabFiltro1['A'+str(count)]=c.montadora
    TabFiltro1['B'+str(count)]=c.modelo
    TabFiltro1['C'+str(count)]=c.tipo
    for y in c.ano:
        t+=str(y)+" "
    TabFiltro1['D'+str(count)]=t
    TabFiltro1['E'+str(count)]=c.capacidade
    TabFiltro1['F'+str(count)]=c.recomendacao

addRotulos(TabFiltro2)
count=1

for c in carrosFiltrados:
    if c.recomendacao!="-":
        t=""
        count+=1
        TabFiltro2['A'+str(count)]=c.montadora
        TabFiltro2['B'+str(count)]=c.modelo
        TabFiltro2['C'+str(count)]=c.tipo
        for y in c.ano:
            t+=str(y)+" "
        TabFiltro2['D'+str(count)]=t
        TabFiltro2['E'+str(count)]=c.capacidade
        TabFiltro2['F'+str(count)]=c.recomendacao    

workbook.save('planilha.xlsx')

#print(planilha['A'+str(num)].value,planilha['B'+str(num)].value,planilha['C'+str(num)].value,planilha['D'+str(num)].value,planilha['E'+str(num)].value,planilha['F'+str(num)].value,planilha['G'+str(num)].value)

#print(planilha.cell(row=1,column=1).value)
#print(consolidada.max_row) #numero maximo de colunas
#print(consolidada.max_column) #numero maximo de colunas