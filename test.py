# Import `os` 
import os
from openpyxl import load_workbook
from lib.Carro import Carro


#variaveis de arquivo
raizArquivo = os.getcwd()+"/"
print(raizArquivo)
planilhasArquivo = raizArquivo+"planilhas/"
saida = raizArquivo+"saida/"
diagramar = "Planilha_Diagramar_Completa_v03.xlsx"
completa = "Planilha AUTOMÓVEIS REV_240918_v2.xlsx"
basica= "basica.xlsx"

workbook= load_workbook(planilhasArquivo+basica)

#Para pegar as worksheets(aba das tabelas) e mostrar a tabela ativa
#print(workbook.sheetnames , workbook.active)

#MONTADORA 	MODELO 	TIPO	 ANO 	CAPACIDADE	RECOMENDACAO_CASTROL
planilha=workbook.get_sheet_by_name("Planilha1")
carros = []
for num in range(681,781):
    carro = Carro(planilha['A'+str(num)].value,planilha['B'+str(num)].value,planilha['C'+str(num)].value,planilha['D'+str(num)].value,planilha['E'+str(num)].value,planilha['F'+str(num)].value)
    carros.append(carro)

    
print(carros)

                

#print(planilha['A'+str(num)].value,planilha['B'+str(num)].value,planilha['C'+str(num)].value,planilha['D'+str(num)].value,planilha['E'+str(num)].value,planilha['F'+str(num)].value,planilha['G'+str(num)].value)

#print(planilha.cell(row=1,column=1).value)
#print(consolidada.max_row) #numero maximo de colunas
#print(consolidada.max_column) #numero maximo de colunas