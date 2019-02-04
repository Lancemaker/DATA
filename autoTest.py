import os
from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl
from lib.Carro import Carro
import re


raizArquivo = os.getcwd()+"/"
print(raizArquivo)
planilhasArquivo = raizArquivo+"planilhas/"
saida = raizArquivo+"saida/"
diagramar = "Planilha_Diagramar_Completa_v03.xlsx"
completa = "Planilha AUTOMÓVEIS REV_240918_v2.xlsx"
basica= "planilha.xlsx"

workbook= load_workbook(planilhasArquivo+basica)

base=workbook.get_sheet_by_name('Planilha1')
planilha=workbook.get_sheet_by_name('Filtro_5')
test=workbook.create_sheet("Testlog")
teste=[]
hit=[]
error=[]
remocoes=[]
e=""
mensagemRemocao=""

def mesToInt(mes):
    meses=["JANEIRO","FEVEREIRO","MARÇO","ABRIL","MAIO","JUNHO","JULHO","AGOSTO","SETEMBRO","OUTUBRO","NOVEMBRO","DEZEMBRO"]
    return meses.index(mes)

for num in range(2,547):
    carroTest=Carro(planilha['A'+str(num)].value,planilha['B'+str(num)].value,planilha['C'+str(num)].value.split(", "),planilha['D'+str(num)].value.split(" "),planilha['E'+str(num)].value.split(", "),planilha['F'+str(num)].value)
    carroTest.ano=carroTest.ano[0]
    if '' in carroTest.ano:
        carroTest.ano.remove('')

    teste.append(carroTest)
    #7834    
for num in range(2,7834):
    carro = Carro(base['A'+str(num)].value,base['B'+str(num)].value,base['C'+str(num)].value,base['D'+str(num)].value,base['E'+str(num)].value,base['F'+str(num)].value)  
    catch = re.findall(r"[-+]?\d*\.\d", carro.modelo)
    if(len(catch)>0):
        carro.tipo = catch[0] 
        carro.modelo = carro.modelo.replace(carro.tipo+" ","")
    else:
        carro.tipo = "NAO ENCONTRADO" 
    if(carro.capacidade != None):
        carro.capacidade=str(carro.capacidade).replace(",",".")   
        if(carro.montadora=="TOYOTA"):  
            carro.ano=str(carro.ano[0]).split(" ")  
    for t in teste:                
        if(carro.recomendacao=="-"):
           remocoes.append(carro)
           mensagemRemocao+="objeto na linha : "+str(num)+" removido deliberadamente por nao ter recomendacao de oleo\n"
           break
        else:
            
            if(carro.montadora == t.montadora):                
                if(t.modelo in carro.modelo)or(t.modelo.replace(" - COM DPF","") in carro.modelo)or(t.modelo.replace(" - SEM DPF","") in carro.modelo):                
                    if(carro.tipo in t.tipo)or(carro.tipo=="NAO ENCONTRADO"):
                        if(t.recomendacao in carro.recomendacao):                            
                            if(carro.capacidade==None)or(carro.capacidade in t.capacidade):                                
                                hit.append(carro)  
                                break
                            else:
                                e="capacidade:"+str(carro.capacidade)+" "+str(t.capacidade) 
                        else:
                            e="recomendação: "+ t.recomendacao +" in "+ carro.recomendacao + " is " + str(t.recomendacao in carro.recomendacao)
                    else:
                        e="tipo: "+ carro.tipo +" in "+ str(t.tipo) + " is " + str(carro.tipo in t.tipo)
                else:
                    e="modelo: "+ t.modelo +" in "+ carro.modelo + " is " + str(t.modelo in carro.modelo)                      
            else:
                e="montadora: "+ carro.montadora +"=="+ t.montadora + " is " + str(carro.montadora == t.montadora)

    if(carro not in hit)and(carro not in remocoes):
        error.append("na linha : "+str(num) + " erro: " +e+ "\n")
print("numero de tentativas: "+str(num-1))
print("numero de acertos: "+str(len(hit)))
print("numero de remocoes: "+str(len(remocoes)))
print("numero de erros: "+str(len(error)))

f= open("ErrorLog.txt","w+")
g= open("Remocoes.txt","w+")
for e in error:
    f.write(e)
f.close()

g.write(mensagemRemocao)
g.close()
