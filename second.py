import os
from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl
from lib.Carro import Carro
import re

#variaveis de arquivo
raizArquivo = os.getcwd()+"/"
print(raizArquivo)
planilhasArquivo = raizArquivo+"planilhas/"
saida = raizArquivo+"saida/"
diagramar = "Planilha_Diagramar_Completa_v03.xlsx"
completa = "Planilha AUTOMÓVEIS REV_240918_v2.xlsx"
basica= "planilha.xlsx"


workbook= load_workbook(planilhasArquivo+basica)

planilha=workbook.get_sheet_by_name('Filtro_2')
TabFiltro3=workbook.create_sheet("Filtro_3")

carros = []
carrosFiltrados=[]

for num in range(2,2923):
    tempString=""
    carro = Carro(planilha['A'+str(num)].value,planilha['B'+str(num)].value,planilha['C'+str(num)].value,planilha['D'+str(num)].value.split(" "),planilha['E'+str(num)].value,planilha['F'+str(num)].value)
    
    carro.ano=carro.ano[0]
    if(carro.ano[-1]==" "):
        carro.ano=carro.ano[:-1]
    #tratando data

    i=0
    while i<len(carro.ano):
        carro.ano[i]=carro.ano[i].replace('[','')
        carro.ano[i]=carro.ano[i].replace(']','')
        carro.ano[i]=carro.ano[i].replace(',','')
        i+=1
    #Regras para renomear os carros:
    temp=carro.modelo.split(" ")
    
    if  ("DPF" in temp):
        for item in temp:
            tempString+=item+" "
        carro.modelo=tempString
       
    else:
        carro.modelo=temp[0]
    if(len(carros)==0):
        carros.append(carro)
    else:
        last=carros[-1]
        if(carro.montadora==last.montadora and carro.modelo==last.modelo and carro.tipo==last.tipo and carro.capacidade==last.capacidade and carro.recomendacao==last.recomendacao):
            x=y=0
            while x<len(carro.ano):                
                while y<len(last.ano):                    
                    if carro.ano[x] not in last.ano:
                        last.ano.append(carro.ano[x])
                        if(not carro.ano[x].isalpha()):
                            last.ano.sort()                        
                    y+=1  
                x+=1
                    
        else:
            carros.append(carro)
for c in carros:
    print(c.montadora,c.modelo,c.tipo,c.ano,c.capacidade,c.recomendacao)
print(len(carros))



#gerando a planilha :D
def addRotulos(tabela):
        tabela['A1']="MONTADORA"
        tabela['B1']="MODELO"
        tabela['C1']="TIPO"
        tabela['D1']="ANO"
        tabela['E1']="CAPACIDADE"
        tabela['F1']="RECOMENDACAO_CASTROL"


addRotulos(TabFiltro3)
count=1
for c in carros:
    t=""
    count+=1
    TabFiltro3['A'+str(count)]=c.montadora
    TabFiltro3['B'+str(count)]=c.modelo
    TabFiltro3['C'+str(count)]=c.tipo
    for y in c.ano:
        t+=str(y)+" "
    TabFiltro3['D'+str(count)]=t
    TabFiltro3['E'+str(count)]=c.capacidade
    TabFiltro3['F'+str(count)]=c.recomendacao
workbook.save('planilha.xlsx')
#print(num,carro.modelo,carro.ano)
#com erro linha 43
#print(num, carrosFiltrados[num-3].montadora, carrosFiltrados[num-3].modelo, carrosFiltrados[num-3].tipo, carrosFiltrados[num-3].ano, carrosFiltrados[num-3].capacidade, carrosFiltrados[num-3].recomendacao)